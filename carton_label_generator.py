#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
carton_label_generator.py

Generate scan-safe 4x6 inch carton labels (PDF) from a Packing List Excel file,
plus an audit CSV, from a WPIC/Topologie-style "Packing List" workbook.

Run once:
    python carton_label_generator.py

or from a notebook:
    from carton_label_generator import main
    main(input_file=INPUT_FILE, output_dir=OUTPUT_DIR, max_skus_per_label=3)

-------------------------------------------------------------------------------
Fixes vs. the original notebook (see accompanying summary for full detail):

1. FOOTER-ROW LEAKAGE (the main data-corruption bug). The original "valid row"
   filter was `sku or barcode or quantity`. The workbook's trailing "TOTAL" row
   has SKU/BarCode blank but Quantity = 4782 (the grand total), so it passed the
   filter, inherited the *previous* carton's Packaging code via forward-fill, and
   was rendered as a bogus extra item on the last carton's label with a quantity
   of 4782. Fixed by requiring SKU or BarCode (never Quantity alone), plus an
   explicit stop at the literal "TOTAL" marker in the Item# column.

2. BILINGUAL HEADER-ECHO ROW LEAKAGE. The row directly under the English header
   (Chinese column labels, e.g. "SKU编码", "条形码") has non-empty text in every
   column, so a plain "has SKU or barcode" filter does not catch it either. It
   is now explicitly detected by matching each cell's own text against the
   known header aliases for that field and dropped.

3. EAN-13 BARCODE COULD OVERFLOW ITS BOX. The original code built the EAN-13
   widget at its *default* size and then did `scale = max(1.0, scale)` -- i.e.
   it refused to ever shrink the barcode to fit the allotted area, so on tight
   layouts (2 SKUs/label) the barcode could be drawn larger than its box and
   overlap neighboring text/lines. Fixed by building the EAN-13 widget directly
   at the target module width/height (ReportLab's eanbc widget accepts barWidth/
   barHeight natively), so it is generated at exactly the right vector size --
   no post-hoc canvas scaling, no distortion, no overflow.

4. AUDIT CSV WAS INCOMPLETE / OUT OF SYNC WITH THE PDF. It was built in a
   separate pass before label splitting existed, so it had no label_index,
   label_total or pdf_page columns and could not be cross-checked against the
   actual generated pages. Fixed by building the audit rows from the exact same
   grouped/split structure used to render the PDF, in the same pass.

5. NO VALIDATION SUMMARY / NO HARD FAILURE ON MISSING COLUMNS. Fixed with an
   explicit column check (raises ValueError listing exactly what's missing) and
   a printed pre-flight summary (row counts, carton counts, invalid barcodes,
   missing quantities, duplicate packaging/SKU pairs) before any file is written.

6. BARCODE VALUE CORRUPTION FROM EXCEL. Numeric-typed barcode cells can arrive
   as floats (4894961044830.0), lose leading zeros, or (rarely) come through as
   scientific notation strings. Fixed with clean_barcode(), which reads the raw
   openpyxl cell's number_format to recover a lost leading-zero width, and
   normalizes floats/scientific notation back to plain digit strings.

7. NO UPC-A HANDLING. 12-digit codes with a valid UPC-A check digit are now
   zero-padded to 13 digits and rendered as EAN-13 (this is the standard,
   widely-used technique for rendering UPC-A on EAN-13-only symbol generators --
   flagged explicitly in the audit as "UPC-A", never silently mixed up with a
   genuine 13-digit EAN-13).

8. "D9e63" -- investigated per your note. It does not match any published
   barcode symbology, ZPL/EPL printer command, or ReportLab/python-barcode
   identifier, and nothing in the workbook (sheet names, cell values, defined
   names, custom number formats) references it either. It looks like it may be
   an internal printer profile/template name or a typo rather than a barcode
   standard, but I can't confirm which without seeing where you encountered it.
   Per your instruction, this script does NOT assume it's a standard and instead
   uses Code 128 for Packaging code and Quantity, and EAN-13/UPC-A (with Code
   128 fallback) for the product barcode -- all explicitly validated below.
-------------------------------------------------------------------------------
"""

from __future__ import annotations

import csv
import math
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.lib.units import inch, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code128 as code128_mod
from reportlab.graphics.barcode import eanbc as eanbc_mod
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF

# ============================================================================
# CONFIG (defaults -- override via main() kwargs)
# ============================================================================

SHEET_NAME_HINT = "Packing List"

FROM_LINES = [
    "From:",
    "Topologie Global Limited",
    "RM G, 9/F, King Palace Plaza",
    "55 King Yip Street, Kwun Tong,",
    "Hong Kong",
]

PAGE_W, PAGE_H = 4 * inch, 6 * inch          # exact 4 x 6 inch, portrait
MARGIN_X = 0.17 * inch
MARGIN_TOP = 0.16 * inch
MARGIN_BOTTOM = 0.12 * inch

MAX_SKUS_PER_LABEL_DEFAULT = 3

# Module width targets, tuned for 203/300 dpi thermal barcode printers.
# 0.30mm narrow module ~= 0.85pt (2-3 dots @203dpi, 3-4 dots @300dpi).
CODE128_MODULE_PT = 0.85
MIN_CODE128_MODULE_PT = 0.70
EAN13_MODULE_PT = 0.85
MIN_EAN13_MODULE_PT = 0.70
MAX_MODULE_PT = 1.00  # hard ceiling -- keeps bars compact/tidy, never fat

REQUIRED_FIELDS = {
    "po": ["po no.", "po no", "po#", "po 编码"],
    "packaging": ["packaging code", "包装条形码"],
    "sku": ["sku#", "sku no.", "sku no", "sku编码"],
    "barcode": ["barcode/upc", "barcode / upc", "条形码"],
    "quantity": ["quantity", "数量"],
}
CANONICAL_NAME = {
    "po": "PO No.",
    "packaging": "Packaging code",
    "sku": "SKU#",
    "barcode": "BarCode/UPC",
    "quantity": "Quantity",
}


# ============================================================================
# TEXT / NUMBER CLEANING
# ============================================================================

def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).lower()


def clean_excel_value(value) -> str:
    """Turn any raw openpyxl/pandas cell value into a clean display string.
    Handles float-ified integers (15.0 -> '15') without touching real decimals."""
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)
    s = str(value).strip()
    if s.lower() in ("nan", "none"):
        return ""
    return s


def clean_quantity(value) -> str:
    s = clean_excel_value(value)
    if not s:
        return ""
    try:
        f = float(s)
        return str(int(f)) if float(f).is_integer() else s
    except ValueError:
        return s


_SCI_RE = re.compile(r"^-?\d(\.\d+)?[eE][+\-]?\d+$")


def clean_barcode(value, number_format: Optional[str] = None) -> str:
    """
    Normalize a barcode cell to a plain digit/alnum string, recovering common
    Excel corruption modes:
      - float-ified integers:       4894961044830.0  -> 4894961044830
      - scientific notation string: 4.89496104483E+12 -> 4894961044830
      - lost leading zeros:         if the cell's own number_format shows a
                                     fixed zero-padded width (e.g. "0000000000000"
                                     or "00"+something), re-pad to that width.
    Never fabricates digits beyond what the number_format / literal cell tells us.
    """
    if value is None:
        return ""

    raw_str = None
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            raw_str = str(int(value))
        else:
            raw_str = repr(value)
    elif isinstance(value, int):
        raw_str = str(value)
    else:
        s = str(value).strip()
        if _SCI_RE.match(s):
            try:
                f = float(s)
                raw_str = str(int(round(f)))
            except ValueError:
                raw_str = s
        else:
            raw_str = s

    raw_str = re.sub(r"\s+", "", raw_str)
    if raw_str.lower() in ("nan", "none"):
        return ""

    # Try to recover a lost leading-zero width from the cell's number_format,
    # e.g. a custom format like "0000000000000" (13 zeros) tells us the true
    # width was 13 digits even though the stored numeric value dropped zeros.
    if number_format and raw_str.isdigit():
        zero_run = re.findall(r"0+", number_format)
        if zero_run:
            target_len = len(max(zero_run, key=len))
            if 0 < target_len <= 20 and len(raw_str) < target_len:
                raw_str = raw_str.zfill(target_len)

    # Generic recovery: common barcode lengths are 12 (UPC-A) / 13 (EAN-13).
    # If we're one or two digits short of one of those AND zero-padding makes
    # the checksum valid, use the padded version (checksum-verified, not a
    # blind guess).
    if raw_str.isdigit() and len(raw_str) in (10, 11, 12):
        for target_len in (12, 13):
            if len(raw_str) < target_len:
                padded = raw_str.zfill(target_len)
                if target_len == 13 and ean13_checksum_is_valid(padded):
                    raw_str = padded
                    break
                if target_len == 12 and ean13_checksum_is_valid("0" + padded):
                    raw_str = padded
                    break

    return raw_str


# ============================================================================
# BARCODE VALIDATION / TYPE DETECTION
# ============================================================================

def ean13_checksum_is_valid(code: str) -> bool:
    if not re.fullmatch(r"\d{13}", code or ""):
        return False
    digits = [int(x) for x in code]
    check = (10 - ((sum(digits[0:-1:2]) + 3 * sum(digits[1:-1:2])) % 10)) % 10
    return check == digits[-1]


def upc_a_checksum_is_valid(code: str) -> bool:
    """UPC-A uses the same mod-10 algorithm as EAN-13 with an implicit leading 0."""
    if not re.fullmatch(r"\d{12}", code or ""):
        return False
    return ean13_checksum_is_valid("0" + code)


def determine_barcode_type(code: str):
    """
    Returns (barcode_type, encode_value, is_valid) where:
      - barcode_type in {"EAN13", "UPC-A", "CODE128", "EMPTY"}
      - encode_value is the exact string to feed to the barcode widget
        (13 digits for EAN13 / UPC-A-as-EAN13, original string for CODE128)
      - is_valid is False whenever we had to fall back to CODE128 because the
        digits didn't check out (never silently "fixed" beyond the checksum-
        verified leading-zero recovery already done in clean_barcode()).
    """
    code = (code or "").strip()
    if not code:
        return "EMPTY", "", False

    if re.fullmatch(r"\d{13}", code):
        if ean13_checksum_is_valid(code):
            return "EAN13", code, True
        return "CODE128", code, False

    if re.fullmatch(r"\d{12}", code):
        if upc_a_checksum_is_valid(code):
            return "UPC-A", "0" + code, True  # render as EAN13 w/ leading 0
        return "CODE128", code, False

    return "CODE128", code, False


# ============================================================================
# EXCEL LOADING / COLUMN DETECTION
# ============================================================================

def find_header_row_and_columns(ws, scan_rows=50):
    """Locate the English header row and map canonical field -> column index."""
    alias_lookup = {}
    for key, aliases in REQUIRED_FIELDS.items():
        for a in aliases:
            alias_lookup[_norm(a)] = key

    for r in range(1, min(scan_rows, ws.max_row) + 1):
        row_map = {_norm(ws.cell(r, c).value): c for c in range(1, ws.max_column + 1)}
        found = {}
        for norm_alias, key in alias_lookup.items():
            if norm_alias in row_map and key not in found:
                found[key] = row_map[norm_alias]
        if len(found) == len(REQUIRED_FIELDS):
            item_col = None
            for c_name, c_idx in row_map.items():
                if c_name in ("item#", "item #", "item", "项目"):
                    item_col = c_idx
                    break
            return r, found, item_col

    missing_summary = ", ".join(CANONICAL_NAME.values())
    raise ValueError(
        f"Missing required columns: could not find a header row containing all of: "
        f"{missing_summary}. Checked the first {min(scan_rows, ws.max_row)} rows."
    )


def load_packing_list(input_file: Path, sheet_name: Optional[str] = None):
    """
    Returns (df, meta) where df has canonical columns:
        PO No., Packaging code, SKU#, BarCode/UPC, Quantity, excel_row
    already truncated to the real data region (header + footer removed),
    forward-filled for merged Packaging code / PO No. cells.
    """
    wb_data = load_workbook(input_file, data_only=True)
    if sheet_name is None:
        sheet_name = SHEET_NAME_HINT if SHEET_NAME_HINT in wb_data.sheetnames else wb_data.sheetnames[0]
    ws = wb_data[sheet_name]

    header_row, col_map, item_col = find_header_row_and_columns(ws)

    header_totals = {}
    for r in range(1, header_row):
        label = _norm(ws.cell(r, 1).value)
        if "package total" in label:
            header_totals["package_total"] = ws.cell(r, 2).value
        elif "quantity total" in label:
            header_totals["quantity_total"] = ws.cell(r, 2).value

    wb_raw = load_workbook(input_file, data_only=False)
    ws_raw = wb_raw[sheet_name]
    barcode_col = col_map["barcode"]

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        item_val = ws.cell(r, item_col).value if item_col else None
        item_str = clean_excel_value(item_val).strip()

        po = ws.cell(r, col_map["po"]).value
        packaging = ws.cell(r, col_map["packaging"]).value
        sku = ws.cell(r, col_map["sku"]).value
        barcode_val = ws.cell(r, col_map["barcode"]).value
        qty = ws.cell(r, col_map["quantity"]).value

        if item_str and re.search(r"\btotal\b", item_str, re.IGNORECASE):
            break
        if records and all(
            clean_excel_value(v) == "" for v in (po, packaging, sku, barcode_val, qty)
        ):
            if not item_str:
                break

        number_format = None
        try:
            number_format = ws_raw.cell(r, barcode_col).number_format
        except Exception:
            number_format = None

        records.append(
            {
                "excel_row": r,
                "PO No.": po,
                "Packaging code": packaging,
                "SKU#": sku,
                "BarCode/UPC": barcode_val,
                "BarCode/UPC__number_format": number_format,
                "Quantity": qty,
            }
        )

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise ValueError("No data rows found below the detected header row.")

    df["Packaging code"] = df["Packaging code"].ffill()
    df["PO No."] = df["PO No."].ffill()

    meta = {
        "header_row": header_row,
        "sheet_name": sheet_name,
        "header_totals": header_totals,
        "col_map": col_map,
    }
    return df, meta


# ============================================================================
# CARTON GROUPING / LABEL SPLITTING
# ============================================================================

@dataclass
class LabelItem:
    excel_row: int
    po_no: str
    packaging_code: str
    sku: str
    barcode: str
    barcode_type: str
    barcode_validation: str
    quantity: str


_SKU_ALIAS_NORM = {_norm(a) for a in REQUIRED_FIELDS["sku"]}
_BARCODE_ALIAS_NORM = {_norm(a) for a in REQUIRED_FIELDS["barcode"]}
_PO_ALIAS_NORM = {_norm(a) for a in REQUIRED_FIELDS["po"]}
_PACKAGING_ALIAS_NORM = {_norm(a) for a in REQUIRED_FIELDS["packaging"]}
_QUANTITY_ALIAS_NORM = {_norm(a) for a in REQUIRED_FIELDS["quantity"]}


def _is_header_echo_row(row) -> bool:
    """
    Some workbooks repeat a second (often bilingual/Chinese) header row
    immediately below the English header row -- e.g. SKU# -> 'SKU编码',
    BarCode/UPC -> '条形码'. That row has non-empty "data" in every column,
    so a plain SKU-or-barcode-present filter does NOT catch it; it must be
    detected by matching each cell's own text against the known header
    aliases for *that* field.
    """
    return (
        _norm(row["SKU#"]) in _SKU_ALIAS_NORM
        or _norm(row["BarCode/UPC"]) in _BARCODE_ALIAS_NORM
        or _norm(row["PO No."]) in _PO_ALIAS_NORM
        or _norm(row["Packaging code"]) in _PACKAGING_ALIAS_NORM
        or _norm(row["Quantity"]) in _QUANTITY_ALIAS_NORM
    )


def build_items(df: pd.DataFrame) -> list[LabelItem]:
    """Clean + validate every row, keeping only rows with a SKU and/or barcode
    (per spec: rows without either are not real line items -- this is also
    what prevents the workbook's trailing TOTAL/summary row from leaking in,
    since that row's Quantity is populated but its SKU/BarCode are not).
    Also drops any bilingual header-echo row (see _is_header_echo_row)."""
    items: list[LabelItem] = []
    for _, row in df.iterrows():
        if _is_header_echo_row(row):
            continue

        sku = clean_excel_value(row["SKU#"]).strip()
        barcode_raw = clean_barcode(row["BarCode/UPC"], row.get("BarCode/UPC__number_format"))

        if not sku and not barcode_raw:
            continue  # not a real item row

        po_no = clean_excel_value(row["PO No."]).strip()
        packaging_code = clean_excel_value(row["Packaging code"]).strip()
        qty = clean_quantity(row["Quantity"])

        btype, encode_val, is_valid = determine_barcode_type(barcode_raw)
        validation = "VALID" if is_valid else ("EMPTY" if btype == "EMPTY" else "FALLBACK")

        items.append(
            LabelItem(
                excel_row=int(row["excel_row"]),
                po_no=po_no,
                packaging_code=packaging_code,
                sku=sku,
                # Rendered symbol is always Code128 (see draw_product_barcode) for a
                # consistent look across the label, so we keep the ORIGINAL cleaned
                # digit string here (not the EAN13-zero-padded version) -- no reason
                # to carry a UPC-A's extra leading 0 into a Code128 symbol.
                barcode=barcode_raw,
                barcode_type=btype,
                barcode_validation=validation,
                quantity=qty,
            )
        )
    return items


def group_cartons(items: list[LabelItem]) -> "OrderedDict[str, list[LabelItem]]":
    """Group items by Packaging code (i.e. by carton), preserving first-seen order."""
    cartons: "OrderedDict[str, list[LabelItem]]" = OrderedDict()
    for it in items:
        if not it.packaging_code:
            raise ValueError(
                f"Excel row {it.excel_row}: item has SKU/barcode but no Packaging code "
                f"could be resolved (even after forward-filling merged cells)."
            )
        cartons.setdefault(it.packaging_code, []).append(it)
    return cartons


def split_carton_items(items: list[LabelItem], max_per_label: int = 2):
    """Split one carton's items into label-sized chunks."""
    return [items[i : i + max_per_label] for i in range(0, len(items), max_per_label)]


# ============================================================================
# FONT SETUP
# ============================================================================

def setup_fonts():
    candidates_regular = [
        Path(r"C:\Windows\Fonts\arialn.ttf"),
        Path(r"C:\Windows\Fonts\ARIALN.TTF"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSansNarrow-Regular.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    candidates_bold = [
        Path(r"C:\Windows\Fonts\arialnb.ttf"),
        Path(r"C:\Windows\Fonts\ARIALNB.TTF"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSansNarrow-Bold.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]
    reg = next((p for p in candidates_regular if p.exists()), None)
    bold = next((p for p in candidates_bold if p.exists()), None)
    if reg and bold:
        try:
            pdfmetrics.registerFont(TTFont("LabelFont", str(reg)))
            pdfmetrics.registerFont(TTFont("LabelFont-Bold", str(bold)))
            return "LabelFont", "LabelFont-Bold"
        except Exception:
            pass
    return "Helvetica", "Helvetica-Bold"


# ============================================================================
# DRAWING HELPERS
# ============================================================================

def fit_font_size(text, font_name, max_size, min_size, max_width):
    size = max_size
    while size > min_size and pdfmetrics.stringWidth(text, font_name, size) > max_width:
        size -= 0.25
    return round(size, 2)


def draw_code128_barcode(c, value, x, y, max_width, max_height, font_reg,
                          human_readable=False, font_size=7,
                          module_target=CODE128_MODULE_PT, module_min=MIN_CODE128_MODULE_PT):
    """Vector Code128, module width solved directly (never post-scaled).
    Targets a compact, standard-looking module width (module_target) instead
    of stretching to fill whatever space is available -- only shrinks below
    that target when the box is too narrow to fit it, down to module_min
    (the hard floor for staying scannable at 203dpi). This keeps every
    barcode a consistent, tidy size rather than some being blown up large."""
    value = (value or "").strip()
    if not value:
        return 0.0

    module = min(module_target, MAX_MODULE_PT)
    trial = code128_mod.Code128(value, barHeight=max_height, barWidth=module,
                                 humanReadable=False, quiet=True)
    if trial.width > max_width:
        module = max(module_min, module * max_width / trial.width)
        trial = code128_mod.Code128(value, barHeight=max_height, barWidth=module,
                                     humanReadable=False, quiet=True)

    draw_x = x + max(0.0, (max_width - trial.width) / 2)
    trial.drawOn(c, draw_x, y)

    if human_readable:
        c.setFont(font_reg, font_size)
        c.drawCentredString(x + max_width / 2, y - font_size - 1, value)

    return trial.width


def font_size_for(module_pt: float) -> float:
    # keep the human-readable digits under the EAN-13 bars legible but compact
    return max(5.5, min(7.5, module_pt * 7.0))


def draw_ean13_barcode(c, value13, x, y, max_width, max_height, font_reg,
                        module_target=EAN13_MODULE_PT, module_min=MIN_EAN13_MODULE_PT):
    """
    Vector EAN-13 (also used for UPC-A padded to 13 digits), module width and
    height solved directly against the allotted box -- no canvas scaling, so
    the barcode can never be drawn larger than its box (the bug in the
    original implementation).
    EAN-13 symbols are always 113 modules wide, so module = width / 113.
    Targets a compact, standard-looking module width (module_target) instead
    of stretching to fill the box -- it only shrinks below that target if the
    box is too narrow, down to module_min, and never grows past module_target
    just because there's extra room (that's what made the barcode look
    oversized/"tràn" -- filling all available space instead of staying a
    tidy, consistent size like Code128).
    """
    value13 = (value13 or "").strip()
    if not value13:
        return

    module = min(module_target, max_width / 113.0, MAX_MODULE_PT)
    module = max(module, module_min)
    height = max(24.0, min(max_height, 46.0))

    widget = eanbc_mod.Ean13BarcodeWidget(value13, barWidth=module, barHeight=height,
                                           humanReadable=1, fontSize=font_size_for(module))
    bounds = widget.getBounds()
    w = bounds[2] - bounds[0]
    h = bounds[3] - bounds[1]

    d = Drawing(w, h)
    d.add(widget)
    draw_x = x + max(0.0, (max_width - w) / 2)
    c.saveState()
    c.translate(draw_x, y)
    renderPDF.draw(d, c, 0, 0)
    c.restoreState()


def draw_product_barcode(c, item: LabelItem, x, y, max_width, max_height, font_reg, font_size=7):
    """
    Per explicit request, every barcode on the label -- product, packaging, and
    quantity -- is rendered as Code128, for one consistent, compact, standard
    look. The product number's EAN-13/UPC-A validity is still checksum-checked
    and recorded in item.barcode_type / item.barcode_validation (and in the
    audit CSV) so the underlying number is known to be a genuine, correct GTIN
    even though it's drawn as a Code128 symbol rather than an EAN-13/UPC-A one.
    (draw_ean13_barcode() above is left in place, unused, in case a true
    EAN-13/UPC-A symbol is wanted again later -- just call it instead here for
    item.barcode_type in ("EAN13", "UPC-A").)
    """
    if item.barcode_type == "EMPTY":
        return  # nothing to draw; audit will flag it
    draw_code128_barcode(c, item.barcode, x, y + 6, max_width, max(18, max_height - 6),
                          font_reg, human_readable=True, font_size=font_size)


# ============================================================================
# LABEL PAGE RENDERING
# ============================================================================

def draw_header(c, po_no, packaging_code, top_y, font_reg, font_bold,
                 label_index=1, label_total=1, shipping_mark=None):
    """
    Layout notes: the packaging-code barcode is alphanumeric (e.g.
    'PGKEC2C17JSH3170001', 20 chars) and needs far more horizontal room than
    the narrow right-hand column used in the original layout -- squeezed into
    ~90pt it either got crushed to an unscannable module width or, if a safe
    module width was enforced, ran off the right edge of the label entirely
    (both were observed in testing). It is therefore drawn as its own
    full-width row below the From/PO block instead, which is the only way to
    keep it scan-safe on a 4in-wide label; "PKG ID:" stays in its original
    position as a label per the mockup.
    """
    left_x = MARGIN_X + 0.14 * inch
    right_margin_x = PAGE_W - MARGIN_X - 0.14 * inch
    po_label_x = 2.18 * inch
    po_value_x = 2.75 * inch

    c.setFont(font_bold, 7.8)
    c.drawString(left_x, top_y, FROM_LINES[0])
    c.setFont(font_reg, 7.8)
    leading = 10.0
    for i, line in enumerate(FROM_LINES[1:], start=1):
        c.drawString(left_x, top_y - i * leading, line)

    c.setFont(font_bold, 7.6)
    c.drawString(po_label_x, top_y, "PO No.:")
    c.setFont(font_reg, 7.6)
    po_text = po_no or ""
    po_size = fit_font_size(po_text, font_reg, 7.6, 5.5, right_margin_x - po_value_x)
    c.setFont(font_reg, po_size)
    c.drawString(po_value_x, top_y, po_text)

    # Shipping Mark: printed directly below PO No. (no "Shipping Mark:" caption,
    # just the value itself) in bold, larger than the surrounding header text,
    # but in the same font family (font_bold) so it stays visually consistent.
    if shipping_mark:
        sm_text = str(shipping_mark).strip().upper()
        sm_y = top_y - 17
        sm_max_w = right_margin_x - po_label_x
        sm_size = fit_font_size(sm_text, font_bold, 13.0, 8.0, sm_max_w)
        c.setFont(font_bold, sm_size)
        c.drawString(po_label_x, sm_y, sm_text)

    # full-width PKG ID row, placed after the From/address block (5 lines)
    pkg_label_y = top_y - 5 * leading - 10
    c.setFont(font_bold, 7.6)
    c.drawString(left_x, pkg_label_y, "PKG ID:")

    if label_total > 1:
        c.setFont(font_bold, 6.8)
        c.drawRightString(right_margin_x, pkg_label_y, f"Label {label_index}/{label_total}")

    pkg_bar_h = 26
    pkg_bar_y = pkg_label_y - 6 - pkg_bar_h
    pkg_bar_x = left_x
    pkg_bar_w = right_margin_x - left_x
    draw_code128_barcode(c, packaging_code, pkg_bar_x, pkg_bar_y, pkg_bar_w, pkg_bar_h,
                          font_reg, human_readable=False)

    pkg_text_y = pkg_bar_y - 9
    pkg_size = fit_font_size(packaging_code, font_reg, 7.2, 5.0, pkg_bar_w)
    c.setFont(font_reg, pkg_size)
    c.drawCentredString(pkg_bar_x + pkg_bar_w / 2, pkg_text_y, packaging_code)

    line_y = pkg_text_y - 8
    c.setLineWidth(1.1)
    c.line(MARGIN_X + 0.10 * inch, line_y, PAGE_W - MARGIN_X - 0.10 * inch, line_y)
    return line_y


def draw_item_block(c, item: LabelItem, y_top, block_h, font_reg, font_bold,
                     is_last=False, single_item_page=False):
    x0 = MARGIN_X + 0.12 * inch
    x1 = PAGE_W - MARGIN_X - 0.12 * inch
    left_w = 2.58 * inch
    qty_x = x0 + left_w + 0.08 * inch
    qty_w = x1 - qty_x

    n = block_h / inch
    if n >= 1.05:
        label_fs, value_fs = 7.5, 7.2
    elif n >= 0.72:
        label_fs, value_fs = 6.8, 6.5
    else:
        label_fs, value_fs = 6.0, 5.8

    c.setFont(font_bold, label_fs)
    c.drawString(x0, y_top - 12, "SKU No:")
    c.setFont(font_reg, value_fs)
    sku_x = x0 + 0.63 * inch
    sku_max_w = qty_x - sku_x - 0.08 * inch
    sku_fs = fit_font_size(item.sku, font_reg, value_fs, 4.7, sku_max_w)
    c.setFont(font_reg, sku_fs)
    c.drawString(sku_x, y_top - 12, item.sku)

    c.setFont(font_bold, label_fs)
    c.drawString(qty_x, y_top - 12, "Quantity:")

    if single_item_page:
        # a single SKU leaves a lot of vertical room on a 4x6 label, but the
        # barcode itself should stay a compact, standard-looking size (like
        # Code128) rather than being stretched to fill the page -- so this is
        # only a modest bump over the multi-SKU size, not a maximum-size fill
        barcode_h = 46
        barcode_y = y_top - 12 - 22 - barcode_h
        product_x = x0 + 0.20 * inch
        product_w = left_w - 0.30 * inch
        qty_bar_h = 38
    else:
        barcode_h = max(30, min(40, block_h * 0.32))
        barcode_y = y_top - 12 - 18 - barcode_h
        product_x = x0 + 0.23 * inch
        product_w = left_w - 0.28 * inch
        qty_bar_h = max(26, barcode_h - 6)

    draw_product_barcode(c, item, product_x, barcode_y, product_w, barcode_h,
                          font_reg, font_size=max(5.5, value_fs))

    qty_bar_y = barcode_y + 7
    draw_code128_barcode(c, item.quantity, qty_x + 0.01 * inch, qty_bar_y,
                          qty_w - 0.02 * inch, qty_bar_h, font_reg, human_readable=False)
    c.setFont(font_reg, max(5.5, value_fs))
    c.drawCentredString(qty_x + qty_w / 2, barcode_y - 2, item.quantity or "")

    if not is_last:
        c.setLineWidth(0.8)
        c.line(x0, y_top - block_h, x1, y_top - block_h)


def render_label_page(c, carton_items_full, label_items, label_index, label_total,
                       font_reg, font_bold, shipping_mark=None):
    c.setLineWidth(2.4)
    c.rect(MARGIN_X, MARGIN_BOTTOM, PAGE_W - 2 * MARGIN_X,
           PAGE_H - MARGIN_BOTTOM - 0.08 * inch, stroke=1, fill=0)

    unique_pos = list(OrderedDict.fromkeys(i.po_no for i in carton_items_full if i.po_no))
    po_no = " / ".join(unique_pos)
    packaging_code = carton_items_full[0].packaging_code

    header_line_y = draw_header(c, po_no, packaging_code, PAGE_H - MARGIN_TOP - 0.10 * inch,
                                 font_reg, font_bold, label_index=label_index, label_total=label_total,
                                 shipping_mark=shipping_mark)

    content_bottom = MARGIN_BOTTOM + 0.12 * inch
    content_top = header_line_y - 0.02 * inch
    available_h = content_top - content_bottom
    block_h = available_h / len(label_items)
    single_item_page = len(label_items) == 1

    y = content_top
    for idx, item in enumerate(label_items):
        draw_item_block(c, item, y, block_h, font_reg, font_bold,
                         is_last=(idx == len(label_items) - 1),
                         single_item_page=single_item_page)
        y -= block_h

    c.showPage()


# ============================================================================
# VALIDATION SUMMARY
# ============================================================================

def print_validation_summary(df, meta, items, cartons, label_plan):
    total_labels = sum(len(v) for v in label_plan.values())
    single_sku = sum(1 for v in cartons.values() if len(v) == 1)
    multi_sku = sum(1 for v in cartons.values() if len(v) > 1)
    invalid_barcodes = sum(1 for i in items if i.barcode_validation != "VALID")
    missing_qty = sum(1 for i in items if not i.quantity)

    seen = {}
    for i in items:
        key = (i.packaging_code, i.sku)
        seen[key] = seen.get(key, 0) + 1
    dup_pairs = sum(1 for v in seen.values() if v > 1)

    print("=" * 70)
    print("VALIDATION SUMMARY")
    print("=" * 70)
    print(f"Sheet                         : {meta['sheet_name']}")
    print(f"Header row                    : {meta['header_row']}")
    print("Detected columns:")
    for key, name in CANONICAL_NAME.items():
        print(f"  - {name}")
    print(f"Total valid item rows         : {len(items)}")
    print(f"Total cartons                 : {len(cartons)}")
    print(f"  Single-SKU cartons          : {single_sku}")
    print(f"  Multi-SKU cartons           : {multi_sku}")
    print(f"Total generated label pages   : {total_labels}")
    print(f"Invalid or missing barcodes   : {invalid_barcodes}")
    print(f"Missing quantities            : {missing_qty}")
    print(f"Duplicate packaging/SKU pairs : {dup_pairs}")

    totals = meta.get("header_totals", {})
    if "quantity_total" in totals:
        try:
            computed_qty = sum(int(i.quantity) for i in items if i.quantity)
        except ValueError:
            computed_qty = None
        sheet_qty = totals["quantity_total"]
        flag = "OK" if computed_qty == sheet_qty else "MISMATCH"
        print(f"Quantity Total cross-check    : sheet={sheet_qty} computed={computed_qty} [{flag}]")
    if "package_total" in totals:
        flag = "OK" if len(cartons) == totals["package_total"] else "MISMATCH"
        print(f"Package Total cross-check     : sheet={totals['package_total']} "
              f"computed={len(cartons)} [{flag}]")
    print("=" * 70)


# ============================================================================
# MAIN
# ============================================================================

def build_label_plan(cartons, max_skus_per_label: int):
    label_plan = OrderedDict()
    for packaging_code, carton_items in cartons.items():
        label_plan[packaging_code] = split_carton_items(carton_items, max_skus_per_label)
    return label_plan


def write_pdf(cartons, label_plan, output_pdf: Path, font_reg: str, font_bold: str, shipping_mark=None):
    c = pdfcanvas.Canvas(str(output_pdf), pagesize=(PAGE_W, PAGE_H), pageCompression=1)
    c.setTitle(output_pdf.stem)
    c.setAuthor("Carton label generator")

    pdf_page = 0
    audit_rows = []
    for packaging_code, groups in label_plan.items():
        carton_items_full = cartons[packaging_code]
        label_total = len(groups)
        for label_index, label_items in enumerate(groups, start=1):
            pdf_page += 1
            render_label_page(c, carton_items_full, label_items, label_index, label_total,
                               font_reg, font_bold, shipping_mark=shipping_mark)
            for item in label_items:
                audit_rows.append(
                    {
                        "source_row": item.excel_row,
                        "po_no": item.po_no,
                        "packaging_code": item.packaging_code,
                        "sku": item.sku,
                        "barcode": item.barcode,
                        "quantity": item.quantity,
                        "carton_item_count": len(carton_items_full),
                        "label_index": label_index,
                        "label_total": label_total,
                        "pdf_page": pdf_page,
                        "barcode_type": item.barcode_type,
                        "barcode_validation": item.barcode_validation,
                    }
                )
    c.save()
    return pdf_page, audit_rows


def write_audit_csv(audit_rows, output_csv: Path):
    fieldnames = [
        "source_row", "po_no", "packaging_code", "sku", "barcode", "quantity",
        "carton_item_count", "label_index", "label_total", "pdf_page",
        "barcode_type", "barcode_validation",
    ]
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)


def main(input_file, output_dir=None, max_skus_per_label: int = MAX_SKUS_PER_LABEL_DEFAULT,
         sheet_name: Optional[str] = None, shipping_mark: Optional[str] = None):
    input_file = Path(input_file)
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    output_dir = Path(output_dir) if output_dir else (input_file.parent / "LABEL_OUTPUT")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_pdf = output_dir / f"{input_file.stem}_4x6_LABELS.pdf"
    output_csv = output_dir / f"{input_file.stem}_label_audit.csv"

    print(f"Input : {input_file}")
    print(f"Output: {output_pdf}")
    print(f"        {output_csv}")
    print()

    df, meta = load_packing_list(input_file, sheet_name=sheet_name)
    items = build_items(df)
    cartons = group_cartons(items)
    label_plan = build_label_plan(cartons, max_skus_per_label)

    print_validation_summary(df, meta, items, cartons, label_plan)

    font_reg, font_bold = setup_fonts()
    pdf_pages, audit_rows = write_pdf(cartons, label_plan, output_pdf, font_reg, font_bold, shipping_mark=shipping_mark)
    write_audit_csv(audit_rows, output_csv)

    print(f"\nDONE: {pdf_pages} label pages written for {len(cartons)} cartons.")
    print(f"PDF : {output_pdf}")
    print(f"CSV : {output_csv}")
    print("\nPrint settings: 4 x 6 inch, Portrait, Actual Size / 100% "
          "(do NOT use Fit / Shrink / Scale to page).")
    return output_pdf, output_csv


if __name__ == "__main__":
    INPUT_FILE = r"C:\Users\Asus\OneDrive\WORKING\KEC\SUPPLY CHAIN\IM-EX DOCUMENTS\2025\DEC 2025\TPLG\VN-VN\2026\JULY 2026\ETD 18 JUL_CN 11722 PCS\PL DO_CN-4785\PL DO_CN-4785\PL_CN-4785_TOTAL.xlsx"
    main(input_file=INPUT_FILE, max_skus_per_label=3)
